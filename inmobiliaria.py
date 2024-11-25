from datetime import *
import openpyxl
import os

# Archivo de la inmobiliaria
archivo_inmobiliaria = 'inmobiliaria.xlsx'

# Función para cargar la hoja de calculo
def cargar_inmobiliaria():
    if os.path.exists(archivo_inmobiliaria):
        workbook = openpyxl.load_workbook(archivo_inmobiliaria)
        sheet = workbook.active
        return workbook,sheet

# Función para cargar la hoja de calculo
def guardar_inmobiliaria(worksheet,fichero):
    # Guardar los cambios en un nuevo archivo Excel
    worksheet.save(fichero)
    print("Libro guardado...")

# Función calcular precio
def calcula_precio(zona,metros,hab,garaje,año_c):
    precio = 0

    if garaje.upper() == "SI":
        g = 1
    else:
        g = 0
    
    antigüedad =  date.today().year - año_c

    # Si está en la zona "A"
    if zona.upper() == "A":
        precio = (metros*1000 + hab*5000 + g*15000) * (1 - antigüedad/100)
    elif zona.upper() == "B":
        precio = (metros*1000 + hab*5000 + g*15000) * (1 - antigüedad/100) * 1.5
    
    return precio

# Funcion buscar inmueble
def buscar_inmueble(hoja,presupuesto):
    # Obtener el índice de la columna "Año Construccion" y "Precio Inmueble"
    encabezados = [cell.value for cell in hoja[1]]
    # En encabezados está esta lista: [Año Construccion,Metros 2,Habitaciones,Garaje,Zona ciudad,Precio Inmueble]

    indice_anio = encabezados.index("Año Construccion")
    indice_metros = encabezados.index("Metros 2")
    indice_hab = encabezados.index("Habitaciones")
    indice_garaje = encabezados.index("Garaje")
    indice_zona = encabezados.index("Zona ciudad")
    indice_precio = encabezados.index("Precio Inmueble")

    cadena = ""
    # Imprimir encabezados
    for elemento in encabezados:
        cadena = cadena + " " + elemento
    print(cadena)

    # Recorrer las filas y recalcular el "Precio Inmueble"
    for row in hoja.iter_rows(min_row=2, values_only=False):
        
        # Obtener los valores de cada fila
        anio_construccion = row[indice_anio].value
        metros = row[indice_metros].value
        habitaciones = row[indice_hab].value
        g = row[indice_garaje].value
        z = row[indice_zona].value
        precio = row[indice_precio].value

        # Inmuebles que interesan por precio
        if precio <= presupuesto:
            print(anio_construccion," ",metros," ",habitaciones," ",g," ",z," ",int(precio))

# Funcion rellenar precios
def rellenar_precios(hoja):
    # Obtener el índice de la columna "Año Construccion" y "Precio Inmueble"
    encabezados = [cell.value for cell in hoja[1]]
    # En encabezados está esta lista: [Año Construccion,Metros 2,Habitaciones,Garaje,Zona ciudad,Precio Inmueble]

    indice_anio = encabezados.index("Año Construccion")
    indice_metros = encabezados.index("Metros 2")
    indice_hab = encabezados.index("Habitaciones")
    indice_garaje = encabezados.index("Garaje")
    indice_zona = encabezados.index("Zona ciudad")
    indice_precio = encabezados.index("Precio Inmueble")

    # Recorrer las filas y recalcular el "Precio Inmueble"
    for row in hoja.iter_rows(min_row=2, values_only=False):
        
        # Obtener los valores de cada fila
        anio_construccion = row[indice_anio].value
        metros = row[indice_metros].value
        habitaciones = row[indice_hab].value
        g = row[indice_garaje].value
        z = row[indice_zona].value

        # Calcula el precio de cada fila
        row[indice_precio].value = calcula_precio(z,int(metros),int(habitaciones),g,anio_construccion)

# Programa principal
def main():
    
    while True:
        inmobiliaria,folio = cargar_inmobiliaria()
        opcion = input("Seleccione opción:\n1.Buscar Inmuebles por presupuesto.\n2.Actualizar precio Inmuebles.\n3.Salir.\n").lower()
        
        if opcion == '1':
            opcion2=int(input("Dime qué presupuesto tienes:"))
            buscar_inmueble(folio,opcion2)

        elif opcion == '2':
            print("Iniciando actuaización...")
            rellenar_precios(folio)
            guardar_inmobiliaria(inmobiliaria,archivo_inmobiliaria)
            print("Actuaización finalizada...")         
        elif opcion == '3':
            print("Saliendo del programa.")
            break
        else:
            print("Opción no válida. Intenta de nuevo.")

if __name__ == "__main__":
    main()