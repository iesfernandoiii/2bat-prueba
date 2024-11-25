import openpyxl
import os

# Archivo de la agenda
archivo_agenda = 'agenda.xlsx'

# Función para cargar la agenda
def cargar_agenda():
    if os.path.exists(archivo_agenda):
        workbook = openpyxl.load_workbook(archivo_agenda)
        sheet = workbook.active
        agenda = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            agenda.append(row)
        return agenda
    else:
        # Si el archivo no existe, creamos uno nuevo con encabezados
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Nombre', 'Apellidos', 'Teléfono', 'Email'])
        workbook.save(archivo_agenda)
        return []

# Función para guardar un nuevo contacto
def guardar_contacto(nombre, apellidos, telefono, email):
    workbook = openpyxl.load_workbook(archivo_agenda)
    sheet = workbook.active
    sheet.append([nombre, apellidos, telefono, email])
    workbook.save(archivo_agenda)

# Función para mostrar la agenda
def mostrar_agenda(agenda):
    if not agenda:
        print("La agenda está vacía.")
    else:
        print("Agenda:")
        for contacto in agenda:
            print(f"Nombre: {contacto[0]}, Apellidos: {contacto[1]}, Teléfono: {contacto[2]}, Email: {contacto[3]}")

# Función para buscar un contacto por nombre
def buscar_contacto(agenda, nombre_buscado):
    for contacto in agenda:
        if contacto[0].lower() == nombre_buscado.lower():
            return contacto
    return None

# Función principal
def main():
    
    
    while True:
        agenda = cargar_agenda()
        opcion = input("¿Deseas añadir un nuevo contacto (a), ver la agenda (v) o buscar un contacto (b)? (Escribe 'salir' para terminar): ").lower()
        
        if opcion == 'a':
            nombre = input("Nombre: ")
            apellidos = input("Apellidos: ")
            telefono = input("Teléfono: ")
            email = input("Email: ")
            guardar_contacto(nombre, apellidos, telefono, email)
            print("Contacto añadido.")

        elif opcion == 'v':
            mostrar_agenda(agenda)

        elif opcion == 'b':
            nombre_buscado = input("Introduce el nombre del contacto a buscar: ")
            contacto = buscar_contacto(agenda, nombre_buscado)
            if contacto:
                print(f"Contacto encontrado: Nombre: {contacto[0]}, Apellidos: {contacto[1]}, Teléfono: {contacto[2]}, Email: {contacto[3]}")
            else:
                print("Contacto no encontrado.")            

        elif opcion == 'salir':
            print("Saliendo del programa.")
            break

        else:
            print("Opción no válida. Intenta de nuevo.")

if __name__ == "__main__":
    main()